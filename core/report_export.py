import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def export_excel(risk_list, save_name="合同审核报告.xlsx"):
    out_path = "output"
    if not os.path.exists(out_path):
        os.makedirs(out_path)
    full_path = os.path.join(out_path, save_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "审核风险清单"
    header = ["风险等级", "审查模块", "问题描述", "整改建议"]
    ws.append(header)

    # 样式
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for risk in risk_list:
        row = [risk["risk_level"], risk["module"], risk["problem"], risk["suggest"]]
        ws.append(row)
        row_idx = ws.max_row
        level = risk["risk_level"]
        if level == "高风险":
            ws[f"A{row_idx}"].fill = red_fill
        elif level == "中风险":
            ws[f"A{row_idx}"].fill = orange_fill
        else:
            ws[f"A{row_idx}"].fill = green_fill

    # 调整列宽
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60

    wb.save(full_path)
    return full_path